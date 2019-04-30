#This app is a tool to quote Rako system based on pre-defined parameters such as number and types of circuits and nature of system
#Complete with PyQt5-based GUI
#The app takes a predefined excel file with a blank quote page and manipulates it based on the parameters selected.
#It then saves the quote in a new excel file, named based on user input
#I'm a complete novice, so please excuse any newbie mistakes you may (and probably will) find in this code


import sys
import math
import openpyxl
from PyQt5 import QtWidgets, uic, QtGui
from PyQt5.QtWidgets import (QApplication, QWizard, QGridLayout, QGroupBox, QPushButton, QRadioButton, QVBoxLayout, QWidget, QLabel, QLineEdit, QComboBox, QHBoxLayout, QInputDialog)


'''class Window(QWidget):
    #this function initializes the window app and sets the layout for the GUI in a grid layout
    def __init__(self, parent=None):
        super(Window, self).__init__(parent)

        grid = QGridLayout()
        grid.addWidget(self.createSystemGroup(), 0, 0)
        grid.addWidget(self.createBridgeGroup(), 1, 0)
        grid.addWidget(self.createCircuitGroup(), 2, 0)
        grid.addWidget(self.createWKeypads(), 0, 1)
        grid.addWidget(self.rf_keys(), 1,1)
        grid.addWidget(self.nfc_keys(), 2,1)
        grid.addWidget(self.generate(), 3,1)
        grid.addWidget(self.misc(), 2,2)
        self.setLayout(grid)
        self.setWindowIcon(QtGui.QIcon('Logo.png'))
        self.setWindowTitle('Rako Quote Bot')
        self.state_system = ''
        self.tc_state = ''

    #This function sets the first group of buttons in the GUI, showing three available options for Rako system (wired, wireless or both)
    def createSystemGroup(self):
        groupSystem = QGroupBox("Is your Rako system wired, wireless or both?")
        wired_sys = QRadioButton("Wired")
        wireless_sys = QRadioButton("Wireless")
        both_sys = QRadioButton("Both")
        vbox = QVBoxLayout()
        vbox.addWidget(wired_sys)
        vbox.addWidget(wireless_sys)
        vbox.addWidget(both_sys)
        groupSystem.setLayout(vbox)
        wireless_sys.toggled.connect(lambda: self.system_clicked(wireless_sys))
        wired_sys.toggled.connect(lambda: self.system_clicked(wired_sys))
        both_sys.toggled.connect(lambda: self.system_clicked(both_sys))
        return groupSystem

    #This function defines a variable (state_system) as the text of the selected radio button from SystemGroup section.
    def system_clicked(self, b):
        if b.isChecked():
            self.state_system = b.text()

    #This function sets the second group of buttons in the GUI
    def createBridgeGroup(self):
        groupBridge = QGroupBox("Time clock required?")
        tc_yes = QRadioButton("Yes")
        tc_no = QRadioButton("No")
        vbox = QVBoxLayout()
        vbox.addWidget(tc_yes)
        vbox.addWidget(tc_no)
        vbox.addStretch()
        groupBridge.setLayout(vbox)
        tc_yes.toggled.connect(lambda: self.time_click(tc_yes))
        tc_no.toggled.connect(lambda: self.time_click(tc_no))
        return groupBridge

    # This function defines a variable (tc_state) as the text of the selected radio button from BridgeGroup section.
    def time_click(self, b):
        if b.isChecked():
            self.tc_state = b.text()

    # This function sets the third group of buttons in the GUI, with text boxes to input quantity of various circuit types.
    def createCircuitGroup(self):
        groupCircuit = QGroupBox("How many circuits in the system?")
        subgrid = QGridLayout()
        wired_lbl = QLabel('Wired')
        wireless_lbl = QLabel('Wireless')
        dim_lbl = QLabel('Dimming:')
        sw_lbl = QLabel('Switching:')
        cub_lbl = QLabel('Curtains/Blinds:')
        self.dim_w_edit = QLineEdit('0')
        self.dim_wl_edit = QLineEdit('0')
        self.dim_wl_edit.setMaximumWidth(30)
        self.dim_w_edit.setMaximumWidth(30)
        self.sw_w_edit = QLineEdit('0')
        self.sw_wl_edit = QLineEdit('0')
        self.sw_wl_edit.setMaximumWidth(30)
        self.sw_w_edit.setMaximumWidth(30)
        self.cub_w_edit = QLineEdit('0')
        self.cub_wl_edit = QLineEdit('0')
        self.cub_wl_edit.setMaximumWidth(30)
        self.cub_w_edit.setMaximumWidth(30)
        vbox = QVBoxLayout()
        vbox.addStretch()
        subgrid.addWidget(wired_lbl, 0,1)
        subgrid.addWidget(wireless_lbl, 0,2)
        subgrid.addWidget(dim_lbl, 1,0)
        subgrid.addWidget(sw_lbl, 2,0)
        subgrid.addWidget(cub_lbl, 3,0)
        subgrid.addWidget(self.dim_w_edit, 1,1)
        subgrid.addWidget(self.dim_wl_edit, 1, 2)
        subgrid.addWidget(self.sw_w_edit,2,1)
        subgrid.addWidget(self.sw_wl_edit,2,2)
        subgrid.addWidget(self.cub_w_edit, 3,1)
        subgrid.addWidget(self.cub_wl_edit,3,2)
        groupCircuit.setLayout(subgrid)
        groupCircuit.setLayout(vbox)
        return groupCircuit

    #This function sets the fourth group of buttons in the GUI, with text boxes to input quantity of various wired keypad types.
    #Additionally there are dropdown menus for keypad plate colour
    def createWKeypads(self):
        groupWK = QGroupBox("How many wired keypads in the system?")
        subgrid = QGridLayout()
        self.w3_qty = QLineEdit('0')
        self.w3_qty.setMaximumWidth(30)
        self.w7_qty = QLineEdit('0')
        self.w7_qty.setMaximumWidth(30)
        self.w10_qty = QLineEdit('0')
        self.w10_qty.setMaximumWidth(30)
        w3_lbl = QLabel('3-button keypads:')
        w7_lbl = QLabel('7-button keypads:')
        w10_lbl = QLabel('10-button keypads:')
        self.w3_plate = QComboBox(self)
        self.w3_plate.addItem('White')
        self.w3_plate.addItem('Black Nickel')
        self.w3_plate.addItem('Polished Brass')
        self.w3_plate.addItem('Mirrored Stainless Steel')
        self.w3_plate.addItem('Brushed Stainless Steel')
        self.w7_plate = QComboBox(self)
        self.w7_plate.addItem('White')
        self.w7_plate.addItem('Black Nickel')
        self.w7_plate.addItem('Polished Brass')
        self.w7_plate.addItem('Mirrored Stainless Steel')
        self.w7_plate.addItem('Brushed Stainless Steel')
        self.w10_plate = QComboBox(self)
        self.w10_plate.addItem('White')
        self.w10_plate.addItem('Black Nickel')
        self.w10_plate.addItem('Polished Brass')
        self.w10_plate.addItem('Mirrored Stainless Steel')
        self.w10_plate.addItem('Brushed Stainless Steel')
        subgrid.addWidget(w3_lbl, 0,0)
        subgrid.addWidget(w7_lbl, 1,0)
        subgrid.addWidget(w10_lbl, 2,0)
        subgrid.addWidget(self.w3_qty, 0,1)
        subgrid.addWidget(self.w7_qty, 1,1)
        subgrid.addWidget(self.w10_qty, 2,1)
        subgrid.addWidget(self.w3_plate,0,2)
        subgrid.addWidget(self.w7_plate,1,2)
        subgrid.addWidget(self.w10_plate,2,2)
        groupWK.setLayout(subgrid)
        return groupWK

    # This function sets the fifth group of buttons in the GUI, with text boxes to input quantity of various wireless keypad types.
    # Additionally there are dropdown menus for keypad plate colour
    def rf_keys(self):
        groupRF = QGroupBox("How many RF wireless keypads in the system?")
        subgrid = QGridLayout()
        self.wl3_qty = QLineEdit('0')
        self.wl3_qty.setMaximumWidth(30)
        self.wl7_qty = QLineEdit('0')
        self.wl7_qty.setMaximumWidth(30)
        self.wl10_qty = QLineEdit('0')
        self.wl10_qty.setMaximumWidth(30)
        w3_lbl = QLabel('3-button keypads:')
        w7_lbl = QLabel('7-button keypads:')
        w10_lbl = QLabel('10-button keypads:')
        self.wl3_plate = QComboBox(self)
        self.wl3_plate.addItem('White')
        self.wl3_plate.addItem('Black Nickel')
        self.wl3_plate.addItem('Polished Brass')
        self.wl3_plate.addItem('Mirrored Stainless Steel')
        self.wl3_plate.addItem('Brushed Stainless Steel')
        self.wl7_plate = QComboBox(self)
        self.wl7_plate.addItem('White')
        self.wl7_plate.addItem('Black Nickel')
        self.wl7_plate.addItem('Polished Brass')
        self.wl7_plate.addItem('Mirrored Stainless Steel')
        self.wl7_plate.addItem('Brushed Stainless Steel')
        self.wl10_plate = QComboBox(self)
        self.wl10_plate.addItem('White')
        self.wl10_plate.addItem('Black Nickel')
        self.wl10_plate.addItem('Polished Brass')
        self.wl10_plate.addItem('Mirrored Stainless Steel')
        self.wl10_plate.addItem('Brushed Stainless Steel')
        subgrid.addWidget(w3_lbl, 0,0)
        subgrid.addWidget(w7_lbl, 1,0)
        subgrid.addWidget(w10_lbl, 2,0)
        subgrid.addWidget(self.wl3_qty, 0,1)
        subgrid.addWidget(self.wl7_qty, 1,1)
        subgrid.addWidget(self.wl10_qty, 2,1)
        subgrid.addWidget(self.wl3_plate,0,2)
        subgrid.addWidget(self.wl7_plate,1,2)
        subgrid.addWidget(self.wl10_plate,2,2)
        groupRF.setLayout(subgrid)
        return groupRF

    # This function sets the sixth group of buttons in the GUI, with text boxes to input quantity of various wireless keypad types.
    # Additionally there are dropdown menus for keypad plate colour
    def nfc_keys(self):
        groupNFC = QGroupBox("How many NFC wireless keypads in the system?")
        subgrid = QGridLayout()
        self.wnf3_qty = QLineEdit('0')
        self.wnf3_qty.setMaximumWidth(30)
        self.wnf7_qty = QLineEdit('0')
        self.wnf7_qty.setMaximumWidth(30)
        self.wnf10_qty = QLineEdit('0')
        self.wnf10_qty.setMaximumWidth(30)
        w7_lbl = QLabel('7-button keypads:')
        w10_lbl = QLabel('10-button keypads:')
        self.wnf7_plate = QComboBox(self)
        self.wnf7_plate.addItem('White')
        self.wnf7_plate.addItem('Black Nickel')
        self.wnf7_plate.addItem('Polished Brass')
        self.wnf7_plate.addItem('Mirrored Stainless Steel')
        self.wnf7_plate.addItem('Brushed Stainless Steel')
        self.wnf10_plate = QComboBox(self)
        self.wnf10_plate.addItem('White')
        self.wnf10_plate.addItem('Black Nickel')
        self.wnf10_plate.addItem('Polished Brass')
        self.wnf10_plate.addItem('Mirrored Stainless Steel')
        self.wnf10_plate.addItem('Brushed Stainless Steel')
        subgrid.addWidget(w7_lbl, 0, 0)
        subgrid.addWidget(w10_lbl, 1, 0)
        subgrid.addWidget(self.wnf7_qty, 0, 1)
        subgrid.addWidget(self.wnf10_qty, 1, 1)
        subgrid.addWidget(self.wnf7_plate, 0, 2)
        subgrid.addWidget(self.wnf10_plate, 1, 2)
        groupNFC.setLayout(subgrid)
        return groupNFC

    #This function builds the "Generate quote" button and links it to the showDialog function
    def generate(self):
        hbox = QHBoxLayout()
        groupcode = QGroupBox()
        btn = QPushButton('Generate Quote')
        btn.setFixedWidth(180)
        hbox.addStretch()
        hbox.addWidget(btn)
        hbox.addStretch()
        groupcode.setLayout(hbox)
        btn.clicked.connect(lambda: self.showDialog())
        return groupcode

    # This function sets the seventh group of buttons in the GUI, to allow user to input qunatity of misc. devices
    def misc(self):
        groupMisc = QGroupBox("Misc. devices")
        subgrid = QGridLayout()
        RAK_STAR = QLabel('RAK-RAK-STAR')
        RAMPI = QLabel('RAMPI')
        TCM = QLabel('TCM')
        leds_lbl = QLabel('Name')
        qty_lbl = QLabel('Qty')
        self.Star_qty = QLineEdit('0')
        self.Star_qty.setMaximumWidth(30)
        self.RAMPI_qty = QLineEdit('0')
        self.RAMPI_qty.setMaximumWidth(30)
        self.TCM_qty = QLineEdit('0')
        self.TCM_qty.setMaximumWidth(30)
        subgrid.addWidget(leds_lbl, 0,0)
        subgrid.addWidget(qty_lbl, 0,1)
        subgrid.addWidget(RAK_STAR, 1,0)
        subgrid.addWidget(RAMPI, 2,0)
        subgrid.addWidget(self.Star_qty, 1,2)
        subgrid.addWidget(self.RAMPI_qty, 2,2)
        subgrid.addWidget(TCM, 3,0)
        subgrid.addWidget(self.TCM_qty, 3,2)
        groupMisc.setLayout(subgrid)
        return groupMisc
'''


#This function manipulates the excel sheet based on user input of wireless circuits
def wireless_modules(self, sheet):
    for i, row in enumerate(sheet.rows, start=1):
        if sheet.cell(row=i, column=1).value == 'RAK-RMT-500':
            sheet.cell(row=i, column=3).value = int(self.dim_wl_edit.text())
        if sheet.cell(row=i, column=1).value == 'RAK-RMS-800':
            sheet.cell(row=i, column=3).value = int(self.sw_wl_edit.text())
        if sheet.cell(row=i, column=1).value == 'RAK-RACUB':
            sheet.cell(row=i, column=3).value = int(self.cub_wl_edit.text())
    x = int(self.dim_wl_edit.text())+int(self.sw_wl_edit.text())+int(self.cub_wl_edit.text())
    rx_link = math.ceil(x / 16)
    for i, row in enumerate(sheet.rows, start=1):
        if sheet.cell(row=i, column=1).value == 'RAK-RX-LINK':
            sheet.cell(row=i, column=3).value = rx_link

'''Defunct functions, please ignore!
def price_check(self):
    rak4t = math.ceil(int(self.dim_w_edit.text()) / 4)
    rak4f = math.ceil(int(self.sw_w_edit.text()) / 4)
    rak4r = math.ceil(int(self.cub_w_edit.text()) / 4)
    rak8 = math.ceil((int(self.dim_w_edit.text()) + int(self.sw_w_edit.text()) + int(self.cub_w_edit.text())) / 8)
    self.t_price = rak4t * 354.2
    self.f_price = rak4f * 339.5
    self.r_price = rak4r * 373.1
    self.rak8_price = rak8 * 249.9
    self.wms_price = int(self.dim_w_edit.text()) * 53.9
    self.wmt_price = int(self.sw_w_edit.text()) * 63.7
    self.wmcub_price = int(self.cub_w_edit.text()) * 63.7
    if (self.t_price+self.f_price+self.r_price) > (self.rak8_price+self.wmcub_price+self.wms_price+self.wmt_price):
        return True
    else:
        return False

def rak4_modules(self, sheet):
    rak4t = math.ceil(int(self.dim_w_edit.text()) / 4)
    rak4f = math.ceil(int(self.sw_w_edit.text()) / 4)
    rak4r = math.ceil(int(self.cub_w_edit.text()) / 4)
    for i, row in enumerate(sheet.rows, start=1):
        if sheet.cell(row=i, column=1).value == 'RAK-RAK4-T':
            sheet.cell(row=i, column=3).value = rak4t
        elif sheet.cell(row=i, column=1).value == 'RAK-RAK4-F':
            sheet.cell(row=i, column=3).value = rak4f
        elif sheet.cell(row=i, column=1).value == 'RAK-RAK4-R':
            sheet.cell(row=i, column=3).value = rak4r
    y = rak4f+rak4t+rak4r
    rak4_link = math.ceil(y / 32)
    for i, row in enumerate(sheet.rows, start=1):
        if sheet.cell(row=i, column=1).value == 'RAK-RAK-LINK':
            sheet.cell(row=i, column=3).value = rak4_link'''

# This function manipulates the excel sheet based on user input of wired circuits
def rak8_modules(self, sheet):
    rak8 = math.ceil((int(self.dim_w_edit.text()) + int(self.sw_w_edit.text()) + int(self.cub_w_edit.text())) / 8)
    for i, row in enumerate(sheet.rows, start=1):
        if sheet.cell(row=i, column=1).value == 'RAK-WMS-600':
            sheet.cell(row=i, column=3).value = int(self.sw_w_edit.text())
        elif sheet.cell(row=i, column=1).value == 'RAK-WMT-400':
            sheet.cell(row=i, column=3).value = int(self.dim_w_edit.text())
        elif sheet.cell(row=i, column=1).value == 'RAK-WM-CUB':
            sheet.cell(row=i, column=3).value = int(self.cub_w_edit.text())
        elif sheet.cell(row=i, column=1).value == 'RAK-RAK8-MB':
            sheet.cell(row=i, column=3).value = rak8
    rak8_link = math.ceil(rak8 / 32)
    for i, row in enumerate(sheet.rows, start=1):
        if sheet.cell(row=i, column=1).value == 'RAK-RAK-LINK':
            sheet.cell(row=i, column=3).value = rak8_link

# This function manipulates the excel sheet based on user input of the radio button options
def wired_bridges(self, sheet):
        if self.tc_state == 'Yes':
            for i, row in enumerate(sheet.rows, start=1):
                if sheet.cell(row=i, column=1).value == 'RAK-WTC-BRIDGE':
                    sheet.cell(row=i, column=3).value = 1
        elif self.tc_state == 'No':
            for i, row in enumerate(sheet.rows, start=1):
                if sheet.cell(row=i, column=1).value == 'RAK-WA-BRIDGE':
                    sheet.cell(row=i, column=3).value = 1

# This function manipulates the excel sheet based on user input of the radio button options
def wireless_bridges(self, sheet):
        if self.tc_state == 'Yes':
            for i, row in enumerate(sheet.rows, start=1):
                if sheet.cell(row=i, column=1).value == 'RAK-RTC-BRIDGE':
                    sheet.cell(row=i, column=3).value = 1
        elif self.tc_state == 'No':
            for i, row in enumerate(sheet.rows, start=1):
                if sheet.cell(row=i, column=1).value == 'RAK-RA-BRIDGE':
                    sheet.cell(row=i, column=3).value = 1

# This function manipulates the excel sheet based on user input of wired keypads
def wired_keypds(self, sheet):
    for i, row in enumerate(sheet.rows, start=1):
        if sheet.cell(row=i, column=1).value == 'RAK-WCM-030':
            sheet.cell(row=i, column=3).value = self.w3_qty.text()
    for i, row in enumerate(sheet.rows, start=1):
        if sheet.cell(row=i, column=1).value == 'RAK-WCM-070':
            sheet.cell(row=i, column=3).value = self.w7_qty.text()
    for i, row in enumerate(sheet.rows, start=1):
        if sheet.cell(row=i, column=1).value == 'RAK-WCM-100':
            sheet.cell(row=i, column=3).value = self.w10_qty.text()

# This function manipulates the excel sheet based on user input of wireless keypads
def rf_keypads(self, sheet):
    for i, row in enumerate(sheet.rows, start=1):
        if sheet.cell(row=i, column=1).value == 'RAK-RCM-030':
            sheet.cell(row=i, column=3).value = self.wl3_qty.text()
    for i, row in enumerate(sheet.rows, start=1):
        if sheet.cell(row=i, column=1).value == 'RAK-RCM-070':
            sheet.cell(row=i, column=3).value = self.wl7_qty.text()
    for i, row in enumerate(sheet.rows, start=1):
        if sheet.cell(row=i, column=1).value == 'RAK-RCM-100':
            sheet.cell(row=i, column=3).value = self.wl10_qty.text()

# This function manipulates the excel sheet based on user input of wireless keypads
def nfc_keypads(self, sheet):
    for i, row in enumerate(sheet.rows, start=1):
        if sheet.cell(row=i, column=1).value == 'RAK-RNC-030':
            sheet.cell(row=i, column=3).value = int(self.wnf3_qty.text())
    for i, row in enumerate(sheet.rows, start=1):
        if sheet.cell(row=i, column=1).value == 'RAK-RNC-070':
            sheet.cell(row=i, column=3).value = self.wnf7_qty.text()
    for i, row in enumerate(sheet.rows, start=1):
        if sheet.cell(row=i, column=1).value == 'RAK-RNC-100':
            sheet.cell(row=i, column=3).value = self.wnf10_qty.text()

# This function manipulates the excel sheet based on user input of keypad plates
def wired_plates(self, sheet):
    if self.w3_plate.currentText() == 'White':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-WLF-030-W':
                sheet.cell(row=i, column=3).value = int(self.w3_qty.text())
    elif self.w3_plate.currentText() == 'Black Nickel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-WLF-030-BN':
                sheet.cell(row=i, column=3).value = int(self.w3_qty.text())
    elif self.w3_plate.currentText() == 'Polished Brass':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-WLF-030-PB':
                sheet.cell(row=i, column=3).value = int(self.w3_qty.text())
    elif self.w3_plate.currentText() == 'Mirrored Stainless Steel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-WLF-030-MSS':
                sheet.cell(row=i, column=3).value = int(self.w3_qty.text())
    elif self.w3_plate.currentText() == 'Brushed Stainless Steel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-WLF-030-SS':
                sheet.cell(row=i, column=3).value = int(self.w3_qty.text())
    if self.w7_plate.currentText() == 'White':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-WLF-070-W':
                sheet.cell(row=i, column=3).value = int(self.w7_qty.text())
    elif self.w7_plate.currentText() == 'Black Nickel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-WLF-070-BN':
                sheet.cell(row=i, column=3).value = int(self.w7_qty.text())
    elif self.w7_plate.currentText() == 'Polished Brass':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-WLF-070-PB':
                sheet.cell(row=i, column=3).value = int(self.w7_qty.text())
    elif self.w7_plate.currentText() == 'Mirrored Stainless Steel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-WLF-070-MSS':
                sheet.cell(row=i, column=3).value = int(self.w7_qty.text())
    elif self.w7_plate.currentText() == 'Brushed Stainless Steel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-WLF-070-SS':
                sheet.cell(row=i, column=3).value = int(self.w7_qty.text())
    if self.w10_plate.currentText() == 'White':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-WLF-100-W':
                sheet.cell(row=i, column=3).value = int(self.w10_qty.text())
    elif self.w10_plate.currentText() == 'Black Nickel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-WLF-100-BN':
                sheet.cell(row=i, column=3).value = int(self.w10_qty.text())
    elif self.w10_plate.currentText() == 'Polished Brass':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-WLF-100-PB':
                sheet.cell(row=i, column=3).value = int(self.w10_qty.text())
    elif self.w10_plate.currentText() == 'Mirrored Stainless Steel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-WLF-100-MSS':
                sheet.cell(row=i, column=3).value = int(self.w10_qty.text())
    elif self.w10_plate.currentText() == 'Brushed Stainless Steel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-WLF-100-SS':
                sheet.cell(row=i, column=3).value = int(self.w10_qty.text())

# This function manipulates the excel sheet based on user input of keypad plates
def wireless_plates(self, sheet):
    self.wl_7 = int(self.wl7_qty.text())
    self.wnf_7 = int(self.wnf7_qty.text())
    self.wl_10 = int(self.wl10_qty.text())
    self.wnf_10 = int(self.wnf10_qty.text())
    if self.wl3_plate.currentText() == 'White':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-030-W':
                sheet.cell(row=i, column=3).value = int(self.wl3_qty.text())
    elif self.wl3_plate.currentText() == 'Black Nickel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-030-BN':
                sheet.cell(row=i, column=3).value = int(self.wl3_qty.text())
    elif self.wl3_plate.currentText() == 'Polished Brass':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-030-PB':
                sheet.cell(row=i, column=3).value = int(self.wl3_qty.text())
    elif self.wl3_plate.currentText() == 'Mirrored Stainless Steel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-030-MSS':
                sheet.cell(row=i, column=3).value = int(self.wl3_qty.text())
    elif self.wl3_plate.currentText() == 'Brushed Stainless Steel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-030-SS':
                sheet.cell(row=i, column=3).value = int(self.wl3_qty.text())
    if self.wl7_plate.currentText() == 'White':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-070-W':
                sheet.cell(row=i, column=3).value = (self.wl_7+self.wnf_7)
    elif self.wl7_plate.currentText() == 'Black Nickel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-070-BN':
                sheet.cell(row=i, column=3).value = (self.wl_7+self.wnf_7)
    elif self.wl7_plate.currentText() == 'Polished Brass':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-070-PB':
                sheet.cell(row=i, column=3).value = (self.wl_7+self.wnf_7)
    elif self.wl7_plate.currentText() == 'Mirrored Stainless Steel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-070-MSS':
                sheet.cell(row=i, column=3).value = (self.wl_7+self.wnf_7)
    elif self.wl7_plate.currentText() == 'Brushed Stainless Steel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-070-SS':
                sheet.cell(row=i, column=3).value = (self.wl_7+self.wnf_7)
    if self.wl10_plate.currentText() == 'White':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-100-W':
                sheet.cell(row=i, column=3).value = (self.wl_10+self.wnf_10)
    elif self.wl10_plate.currentText() == 'Black Nickel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-100-BN':
                sheet.cell(row=i, column=3).value = (self.wl_10+self.wnf_10)
    elif self.wl10_plate.currentText() == 'Polished Brass':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-100-PB':
                sheet.cell(row=i, column=3).value = (self.wl_10+self.wnf_10)
    elif self.wl10_plate.currentText() == 'Mirrored Stainless Steel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-100-MSS':
                sheet.cell(row=i, column=3).value = (self.wl_10+self.wnf_10)
    elif self.wl10_plate.currentText() == 'Brushed Stainless Steel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-100-SS':
                sheet.cell(row=i, column=3).value = (self.wl_10+self.wnf_10)

# This function manipulates the excel sheet based on user input of keypad plates
def nfc_plates(self, sheet):
    if self.wnf7_plate.currentText() == 'White':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-070-W':
                sheet.cell(row=i, column=3).value = (self.wl_7+self.wnf_7)
    elif self.wnf7_plate.currentText() == 'Black Nickel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-070-BN':
                sheet.cell(row=i, column=3).value = (self.wl_7+self.wnf_7)
    elif self.wnf7_plate.currentText() == 'Polished Brass':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-070-PB':
                sheet.cell(row=i, column=3).value = (self.wl_7+self.wnf_7)
    elif self.wnf7_plate.currentText() == 'Mirrored Stainless Steel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-070-MSS':
                sheet.cell(row=i, column=3).value = (self.wl_7+self.wnf_7)
    elif self.wnf7_plate.currentText() == 'Brushed Stainless Steel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-070-SS':
                sheet.cell(row=i, column=3).value = (self.wl_7+self.wnf_7)
    if self.wnf10_plate.currentText() == 'White':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-100-W':
                sheet.cell(row=i, column=3).value = (self.wl_10+self.wnf_10)
    elif self.wnf10_plate.currentText() == 'Black Nickel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-100-BN':
                sheet.cell(row=i, column=3).value = (self.wl_10+self.wnf_10)
    elif self.wnf10_plate.currentText() == 'Polished Brass':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-100-PB':
                sheet.cell(row=i, column=3).value = (self.wl_10+self.wnf_10)
    elif self.wnf10_plate.currentText() == 'Mirrored Stainless Steel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-100-MSS':
                sheet.cell(row=i, column=3).value = (self.wl_10+self.wnf_10)
    elif self.wnf10_plate.currentText() == 'Brushed Stainless Steel':
        for i, row in enumerate(sheet.rows, start=1):
            if sheet.cell(row=i, column=1).value == 'RAK-RLF-100-SS':
                sheet.cell(row=i, column=3).value = (self.wl_10+self.wnf_10)

# This function manipulates the excel sheet based on user input of misc. devices
def misc_items(self, sheet):
    for i, row in enumerate(sheet.rows, start=1):
        if sheet.cell(row=i, column=1).value == 'RAK-RAK-STAR':
            sheet.cell(row=i, column=3).value = self.Star_qty.text()
    for i, row in enumerate(sheet.rows, start=1):
        if sheet.cell(row=i, column=1).value == 'RAK-RAMPI':
            sheet.cell(row=i, column=3).value = self.RAMPI_qty.text()
    for i, row in enumerate(sheet.rows, start=1):
        if sheet.cell(row=i, column=1).value == 'RAK-TCM':
            sheet.cell(row=i, column=3).value = self.TCM_qty.text()


#This function opens up pop-up message when pressing the "Generate Quote" icon, and operates the "gen_quote" function when OK is pressed
def showDialog(self):
    self.book = openpyxl.load_workbook('Rak_Prices.xlsx')
    self.sheet = self.book.active
    self.text, ok = QInputDialog.getText(self, 'Rako Quote Bot','Which company are you quoting for?')
    if ok:
        self.gen_quote(self.text, self.sheet)

#This function puts all the other excel manipulation functions together in basic if/elif logic and user input.
#The function then saves the manipulated excel sheet based on user input name
def gen_quote(self, text, quote_sheet):
    if self.state_system == 'Wired':
        self.wired_bridges(quote_sheet)
        self.rak8_modules(quote_sheet)
        self.wired_keypds(quote_sheet)
        self.wired_plates(quote_sheet)
        self.misc_items(quote_sheet)
    elif self.state_system == 'Both':
        self.rak8_modules(quote_sheet)
        self.wireless_modules(quote_sheet)
        self.wired_keypds(quote_sheet)
        self.rf_keypads(quote_sheet)
        self.nfc_keypads(quote_sheet)
        self.wired_plates(quote_sheet)
        self.wireless_plates(quote_sheet)
        self.nfc_plates(quote_sheet)
        self.misc_items(quote_sheet)
    elif self.state_system == 'Wireless':
        self.wireless_bridges(quote_sheet)
        self.wireless_modules(quote_sheet)
        self.rf_keypads(quote_sheet)
        self.nfc_keypads(quote_sheet)
        self.wireless_plates(quote_sheet)
        self.nfc_plates(quote_sheet)
        self.misc_items(quote_sheet)
    name = str(self.text)
    name = name + ' Rako Quote.xlsx'
    self.book.save(name)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    dlg = uic.loadUi("Rako_quote_UI-v2.ui")
    dlg.show()
    btn = dlg.FinishButton
    btn.clicked.connect(lambda: self.showDialog())
    app.exec()
    sys.exit(app.exec_())



